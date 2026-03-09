"""
폐기불량 관리시스템 - Python Flask 버전
내부 네트워크용 (인터넷 불필요)
SQLite DB 기반
"""

from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import sqlite3
import os
import uuid
import gzip
import io

app = Flask(__name__)

# 설정
DB_FILE = 'scrap_data.db'
ADMIN_PASSWORD = '5555'

# 마스터 데이터 테이블 설정 (API 경로명 → DB 매핑)
TABLE_CONFIG = {
    'Depart': {
        'table': 'departments',
        'headers': ['부서'],
        'columns': ['name'],
    },
    'Process': {
        'table': 'processes',
        'headers': ['Part', '공정'],
        'columns': ['part', 'name'],
    },
    'machine': {
        'table': 'machines',
        'headers': ['Part', '공정', '설비명'],
        'columns': ['part', 'process', 'name'],
    },
    'person': {
        'table': 'persons',
        'headers': ['폐기자'],
        'columns': ['name'],
    },
    '1Part_TMNO': {
        'table': 'tmnos',
        'headers': ['TM-NO', '품명', '단위중량', '성형', '소결', '후처리'],
        'columns': ['tmno', 'product_name', 'unit_weight', 'forming', 'sintering', 'post_processing'],
        'filter': {'part_type': '1Part'},
        'insert_extra': {'part_type': '1Part'},
    },
    '2Part_TMNO': {
        'table': 'tmnos',
        'headers': ['TM-NO', '품명', '단위중량', '성형', '소결', '후처리'],
        'columns': ['tmno', 'product_name', 'unit_weight', 'forming', 'sintering', 'post_processing'],
        'filter': {'part_type': '2Part'},
        'insert_extra': {'part_type': '2Part'},
    },
    'scrap_name': {
        'table': 'scrap_reasons',
        'headers': ['폐기사유'],
        'columns': ['name'],
    },
}


# ==================== gzip 압축 (Python 내장 모듈) ====================

@app.after_request
def compress_response(response):
    """JSON 응답을 gzip 압축 (Python 내장 gzip 모듈 사용)"""
    if (response.status_code < 200 or response.status_code >= 300
            or 'Content-Encoding' in response.headers
            or not response.content_type.startswith('application/json')):
        return response

    accept_encoding = request.headers.get('Accept-Encoding', '')
    if 'gzip' not in accept_encoding:
        return response

    data = response.get_data()
    if len(data) < 500:
        return response

    buf = io.BytesIO()
    with gzip.GzipFile(fileobj=buf, mode='wb', compresslevel=6) as f:
        f.write(data)

    response.set_data(buf.getvalue())
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = len(response.get_data())
    response.headers['Vary'] = 'Accept-Encoding'
    return response


# ==================== DB 함수 ====================

def get_db():
    """DB 연결 (WAL 모드로 동시 접근 개선)"""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    """DB 테이블 초기화"""
    conn = get_db()
    cur = conn.cursor()

    cur.execute('''
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS processes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part TEXT NOT NULL,
            name TEXT NOT NULL,
            UNIQUE(part, name)
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS machines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part TEXT NOT NULL,
            process TEXT NOT NULL,
            name TEXT NOT NULL,
            UNIQUE(part, process, name)
        )
    ''')

    # persons 테이블 마이그레이션 (이전: part/process/department/name → 신규: name만)
    try:
        cur.execute("SELECT part FROM persons LIMIT 1")
        # 이전 스키마 존재 → 고유 이름만 추출 후 테이블 재생성
        cur.execute("SELECT DISTINCT name FROM persons WHERE name != '' ORDER BY name")
        old_names = [row['name'] for row in cur.fetchall()]
        cur.execute("DROP TABLE persons")
        cur.execute('''
            CREATE TABLE persons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            )
        ''')
        for n in old_names:
            cur.execute("INSERT OR IGNORE INTO persons (name) VALUES (?)", (n,))
        print(f"persons 테이블 마이그레이션 완료: {len(old_names)}명")
    except sqlite3.OperationalError:
        cur.execute('''
            CREATE TABLE IF NOT EXISTS persons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            )
        ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS tmnos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_type TEXT NOT NULL,
            tmno TEXT NOT NULL,
            product_name TEXT DEFAULT '',
            unit_weight REAL DEFAULT 0,
            forming TEXT DEFAULT '',
            sintering TEXT DEFAULT '',
            post_processing TEXT DEFAULT '',
            UNIQUE(part_type, tmno)
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS scrap_reasons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    ''')

    cur.execute('''
        CREATE TABLE IF NOT EXISTS scrap_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            unique_id TEXT UNIQUE NOT NULL,
            date TIMESTAMP NOT NULL,
            part TEXT NOT NULL,
            department TEXT,
            process TEXT,
            machine TEXT,
            person TEXT,
            tmno TEXT,
            product_name TEXT,
            scrap_reason TEXT NOT NULL,
            quantity REAL DEFAULT 0,
            weight REAL DEFAULT 0,
            remark TEXT,
            defect_category TEXT DEFAULT '',
            defect_process TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 기존 테이블에 새 컬럼 추가 (마이그레이션)
    try:
        cur.execute("SELECT defect_category FROM scrap_data LIMIT 1")
    except sqlite3.OperationalError:
        cur.execute("ALTER TABLE scrap_data ADD COLUMN defect_category TEXT DEFAULT ''")
        cur.execute("ALTER TABLE scrap_data ADD COLUMN defect_process TEXT DEFAULT ''")
        print("scrap_data 테이블에 defect_category, defect_process 컬럼 추가 완료")

    # 인덱스
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_date ON scrap_data(date)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_part ON scrap_data(part)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_dept ON scrap_data(department)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_process ON scrap_data(process)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_reason ON scrap_data(scrap_reason)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scrap_tmno ON scrap_data(tmno)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_tmno_part ON tmnos(part_type)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_processes_part ON processes(part)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_machines_part_process ON machines(part, process)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_tmnos_part_tmno ON tmnos(part_type, tmno)')

    conn.commit()
    conn.close()


def migrate_from_excel():
    """기존 Excel 데이터를 DB로 마이그레이션 (최초 1회)"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return

    if not os.path.exists('scrap_data.xlsx'):
        return

    conn = get_db()
    cur = conn.cursor()

    # 이미 마이그레이션됐는지 확인
    cur.execute("SELECT COUNT(*) as cnt FROM departments")
    if cur.fetchone()['cnt'] > 0:
        conn.close()
        return

    print("Excel 데이터 마이그레이션 시작...")
    wb = load_workbook('scrap_data.xlsx')

    # 부서
    for row in wb['Depart'].iter_rows(min_row=2, values_only=True):
        if row[0]:
            cur.execute("INSERT OR IGNORE INTO departments (name) VALUES (?)", (str(row[0]),))

    # 공정
    for row in wb['Process'].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            cur.execute("INSERT OR IGNORE INTO processes (part, name) VALUES (?, ?)",
                       (str(row[0]), str(row[1])))

    # 설비
    for row in wb['machine'].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2]:
            cur.execute("INSERT OR IGNORE INTO machines (part, process, name) VALUES (?, ?, ?)",
                       (str(row[0]), str(row[1]), str(row[2])))

    # 폐기자 (이름만 추출)
    for row in wb['person'].iter_rows(min_row=2, values_only=True):
        if row[3]:
            cur.execute("INSERT OR IGNORE INTO persons (name) VALUES (?)", (str(row[3]),))

    # TM-NO (1Part + 2Part)
    for part_type, sheet_name in [('1Part', '1Part_TMNO'), ('2Part', '2Part_TMNO')]:
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if row[0]:
                cur.execute("""INSERT OR IGNORE INTO tmnos
                    (part_type, tmno, product_name, unit_weight, forming, sintering, post_processing)
                    VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (part_type, str(row[0]), str(row[1] or ''),
                     float(str(row[2])) if row[2] else 0,
                     str(row[3] or ''), str(row[4] or ''), str(row[5] or '')))

    # 폐기사유
    for row in wb['scrap_name'].iter_rows(min_row=2, values_only=True):
        if row[0]:
            cur.execute("INSERT OR IGNORE INTO scrap_reasons (name) VALUES (?)", (str(row[0]),))

    # 기존 폐기 기록 (scrap_data에 이미 있으면 건너뜀)
    for row in wb['Data'].iter_rows(min_row=2, values_only=True):
        if row[0]:
            cur.execute("""INSERT OR IGNORE INTO scrap_data
                (unique_id, date, part, department, process, machine, person,
                 tmno, product_name, scrap_reason, quantity, weight, remark)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (str(row[0]), str(row[1]), str(row[2] or ''), str(row[3] or ''),
                 str(row[4] or ''), str(row[5] or ''), str(row[6] or ''),
                 str(row[7]) if row[7] else '', str(row[8] or ''), str(row[9] or ''),
                 float(row[10] or 0), float(row[11] or 0), str(row[12] or '')))

    conn.commit()
    conn.close()
    wb.close()
    print("Excel 데이터 마이그레이션 완료!")


def generate_unique_id():
    """고유 ID 생성 (UUID 기반, 충돌 방지)"""
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    return f'SC{timestamp}{uuid.uuid4().hex[:6].upper()}'


# ==================== 라우트 ====================

@app.route('/')
def index():
    return render_template('index.html')


# ==================== 마스터 데이터 일괄 로드 (앱 시작 시 1회) ====================

@app.route('/api/init_data', methods=['GET'])
def get_init_data():
    """부서/인원/불량유형을 한번에 반환 (네트워크 왕복 최소화)"""
    conn = get_db()
    departments = [r['name'] for r in conn.execute(
        "SELECT name FROM departments ORDER BY name").fetchall()]
    persons = [r['name'] for r in conn.execute(
        "SELECT name FROM persons ORDER BY name").fetchall()]
    scrap_reasons = [r['name'] for r in conn.execute(
        "SELECT name FROM scrap_reasons ORDER BY id").fetchall()]
    conn.close()
    return jsonify({
        'departments': departments,
        'persons': persons,
        'scrapReasons': scrap_reasons
    })


# ==================== 데이터 조회 API ====================

@app.route('/api/departments', methods=['GET'])
def get_departments():
    conn = get_db()
    rows = conn.execute("SELECT name FROM departments ORDER BY name").fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])


@app.route('/api/processes', methods=['GET'])
def get_processes():
    part = request.args.get('part', '')
    conn = get_db()
    if part:
        rows = conn.execute(
            "SELECT DISTINCT name FROM processes WHERE part = ? ORDER BY id", (part,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT DISTINCT name FROM processes ORDER BY id").fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])


@app.route('/api/machines', methods=['GET'])
def get_machines():
    part = request.args.get('part', '')
    process = request.args.get('process', '')
    conn = get_db()
    rows = conn.execute(
        "SELECT name FROM machines WHERE part = ? AND process = ? ORDER BY name",
        (part, process)
    ).fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])


@app.route('/api/persons', methods=['GET'])
def get_persons():
    conn = get_db()
    rows = conn.execute("SELECT name FROM persons ORDER BY name").fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])


@app.route('/api/tmnos', methods=['GET'])
def get_tmnos():
    part = request.args.get('part', '1Part')
    process = request.args.get('process', '')

    if process not in ['성형', '소결']:
        mapped_process = '후처리'
    else:
        mapped_process = process

    process_col_map = {'성형': 'forming', '소결': 'sintering', '후처리': 'post_processing'}
    col = process_col_map.get(mapped_process)

    conn = get_db()
    if col:
        rows = conn.execute(
            f"SELECT tmno FROM tmnos WHERE part_type = ? AND {col} = 'y' ORDER BY tmno",
            (part,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT tmno FROM tmnos WHERE part_type = ? ORDER BY tmno", (part,)
        ).fetchall()
    conn.close()
    return jsonify([r['tmno'] for r in rows])


@app.route('/api/tmno_info', methods=['GET'])
def get_tmno_info():
    part = request.args.get('part', '1Part')
    tmno = request.args.get('tmno', '')

    conn = get_db()
    row = conn.execute(
        "SELECT tmno, product_name, unit_weight FROM tmnos WHERE part_type = ? AND tmno = ?",
        (part, tmno)
    ).fetchone()
    conn.close()

    if row:
        return jsonify({
            'tmno': row['tmno'],
            'productName': row['product_name'] or '',
            'unitWeight': row['unit_weight'] or 0
        })
    return jsonify({'tmno': tmno, 'productName': '', 'unitWeight': 0})


@app.route('/api/scrap_reasons', methods=['GET'])
def get_scrap_reasons():
    conn = get_db()
    rows = conn.execute("SELECT name FROM scrap_reasons ORDER BY id").fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])


# ==================== 데이터 저장 API ====================

@app.route('/api/save_scrap', methods=['POST'])
def save_scrap():
    data = request.json
    unique_id = generate_unique_id()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    try:
        conn.execute('''
            INSERT INTO scrap_data
            (unique_id, date, part, department, process, machine, person,
             tmno, product_name, scrap_reason, quantity, weight, remark,
             defect_category, defect_process)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            unique_id, now,
            data.get('part', ''),
            data.get('department', ''),
            data.get('process', ''),
            data.get('machine', ''),
            data.get('person', ''),
            str(data.get('tmno', '')),
            data.get('productName', ''),
            data.get('scrapReason', ''),
            data.get('quantity', 0),
            data.get('weight', 0),
            data.get('remark', ''),
            data.get('defectCategory', ''),
            data.get('defectProcess', '')
        ))
        conn.commit()
        return jsonify({'success': True, 'message': '저장되었습니다.', 'id': unique_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': f'저장 실패: {str(e)}'})
    finally:
        conn.close()


# ==================== 마스터 데이터 관리 API ====================

@app.route('/api/master_data/<sheet_name>', methods=['GET'])
def get_master_data(sheet_name):
    if sheet_name not in TABLE_CONFIG:
        return jsonify({'success': False, 'message': '테이블을 찾을 수 없습니다.'})

    config = TABLE_CONFIG[sheet_name]
    table = config['table']
    columns = config['columns']

    conn = get_db()

    where = ""
    params = []
    if 'filter' in config:
        conditions = [f"{k} = ?" for k in config['filter']]
        where = "WHERE " + " AND ".join(conditions)
        params = list(config['filter'].values())

    cols = ', '.join(columns)
    rows = conn.execute(f"SELECT id, {cols} FROM {table} {where} ORDER BY id", params).fetchall()
    conn.close()

    result_rows = []
    for row in rows:
        result_rows.append({
            'rowIndex': row['id'],
            'data': [row[col] for col in columns]
        })

    return jsonify({'success': True, 'headers': config['headers'], 'rows': result_rows})


@app.route('/api/master_data/<sheet_name>', methods=['POST'])
def add_master_data(sheet_name):
    if sheet_name not in TABLE_CONFIG:
        return jsonify({'success': False, 'message': '테이블을 찾을 수 없습니다.'})

    config = TABLE_CONFIG[sheet_name]
    table = config['table']
    columns = list(config['columns'])
    values = list(request.json.get('data', []))

    # TMNO 등 추가 필드 (part_type)
    if 'insert_extra' in config:
        for k, v in config['insert_extra'].items():
            columns.insert(0, k)
            values.insert(0, v)

    conn = get_db()
    try:
        placeholders = ', '.join(['?'] * len(columns))
        col_names = ', '.join(columns)
        conn.execute(f"INSERT INTO {table} ({col_names}) VALUES ({placeholders})", values)
        conn.commit()
        return jsonify({'success': True, 'message': '추가되었습니다.'})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': '이미 존재하는 데이터입니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'추가 실패: {str(e)}'})
    finally:
        conn.close()


@app.route('/api/master_data/<sheet_name>/<int:row_id>', methods=['PUT'])
def update_master_data(sheet_name, row_id):
    password = request.json.get('password', '')
    if password != ADMIN_PASSWORD:
        return jsonify({'success': False, 'message': '비밀번호가 올바르지 않습니다.'})

    # 폐기 기록 수정
    if sheet_name == 'Data':
        data = request.json.get('data', [])
        conn = get_db()
        try:
            conn.execute("""
                UPDATE scrap_data SET unique_id=?, date=?, part=?, department=?,
                process=?, machine=?, person=?, tmno=?, product_name=?,
                scrap_reason=?, quantity=?, weight=?, remark=?,
                defect_category=?, defect_process=? WHERE id=?
            """, data + [row_id])
            conn.commit()
            return jsonify({'success': True, 'message': '수정되었습니다.'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'수정 실패: {str(e)}'})
        finally:
            conn.close()

    if sheet_name not in TABLE_CONFIG:
        return jsonify({'success': False, 'message': '테이블을 찾을 수 없습니다.'})

    config = TABLE_CONFIG[sheet_name]
    table = config['table']
    columns = config['columns']
    data = request.json.get('data', [])

    conn = get_db()
    try:
        set_clause = ', '.join([f"{col} = ?" for col in columns])
        conn.execute(f"UPDATE {table} SET {set_clause} WHERE id = ?", list(data) + [row_id])
        conn.commit()
        return jsonify({'success': True, 'message': '수정되었습니다.'})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': '이미 존재하는 데이터입니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'수정 실패: {str(e)}'})
    finally:
        conn.close()


@app.route('/api/master_data/<sheet_name>/<int:row_id>', methods=['DELETE'])
def delete_master_data(sheet_name, row_id):
    password = request.json.get('password', '')
    if password != ADMIN_PASSWORD:
        return jsonify({'success': False, 'message': '비밀번호가 올바르지 않습니다.'})

    # 폐기 기록 삭제
    if sheet_name == 'Data':
        conn = get_db()
        try:
            conn.execute("DELETE FROM scrap_data WHERE id = ?", (row_id,))
            conn.commit()
            return jsonify({'success': True, 'message': '삭제되었습니다.'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'삭제 실패: {str(e)}'})
        finally:
            conn.close()

    if sheet_name not in TABLE_CONFIG:
        return jsonify({'success': False, 'message': '테이블을 찾을 수 없습니다.'})

    config = TABLE_CONFIG[sheet_name]
    conn = get_db()
    try:
        conn.execute(f"DELETE FROM {config['table']} WHERE id = ?", (row_id,))
        conn.commit()
        return jsonify({'success': True, 'message': '삭제되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'삭제 실패: {str(e)}'})
    finally:
        conn.close()


# ==================== 폐기 기록 조회 API ====================

@app.route('/api/scrap_records', methods=['GET'])
def get_scrap_records():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    per_page = min(per_page, 500)
    offset = (page - 1) * per_page

    # 필터 파라미터
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    part_filter = request.args.get('part', '')
    reason_filter = request.args.get('reason_type', '')  # main / etc / (빈값=전체)
    defect_filter = request.args.get('defect', '')  # 해당 / empty / (빈값=전체)

    where_clauses = []
    params = []

    if start_date:
        where_clauses.append("date >= ?")
        params.append(start_date)
    if end_date:
        where_clauses.append("date <= ?")
        params.append(end_date + ' 23:59:59')
    if part_filter:
        where_clauses.append("part = ?")
        params.append(part_filter)
    if reason_filter == 'main':
        where_clauses.append("scrap_reason IN ('공정불량', '셋팅불량')")
    elif reason_filter == 'etc':
        where_clauses.append("scrap_reason NOT IN ('공정불량', '셋팅불량')")
    if defect_filter == '해당':
        where_clauses.append("defect_category = '해당'")
    elif defect_filter == 'empty':
        where_clauses.append("(defect_category IS NULL OR defect_category = '')")

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    conn = get_db()
    total = conn.execute(f"SELECT COUNT(*) as cnt FROM scrap_data{where_sql}", params).fetchone()['cnt']
    rows = conn.execute(f"""
        SELECT id, unique_id, date, part, department, process, machine, person,
               tmno, product_name, scrap_reason, quantity, weight, remark,
               defect_category, defect_process
        FROM scrap_data{where_sql} ORDER BY id DESC LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()
    conn.close()

    headers = ['ID', '날짜', 'Part', '부서', '공정', '설비명', '폐기자',
               'TM-NO', '품명', '폐기사유', '수량', '중량(kg)', '비고',
               '불량해당', '해당공정']
    result_rows = []
    for row in rows:
        result_rows.append({
            'rowIndex': row['id'],
            'data': [row['unique_id'], row['date'], row['part'], row['department'],
                    row['process'], row['machine'], row['person'], row['tmno'],
                    row['product_name'], row['scrap_reason'], row['quantity'],
                    row['weight'], row['remark'],
                    row['defect_category'] or '', row['defect_process'] or '']
        })

    return jsonify({
        'success': True, 'headers': headers, 'rows': result_rows,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total,
            'total_pages': max(1, (total + per_page - 1) // per_page)
        }
    })


# ==================== 기타 API ====================

@app.route('/api/verify_password', methods=['POST'])
def verify_password():
    password = request.json.get('password', '')
    return jsonify(password == ADMIN_PASSWORD)


@app.route('/api/simple_list/<sheet_name>', methods=['GET'])
def get_simple_list(sheet_name):
    if sheet_name not in TABLE_CONFIG:
        return jsonify([])

    config = TABLE_CONFIG[sheet_name]
    table = config['table']
    first_col = config['columns'][0]

    conn = get_db()
    where = ""
    params = []
    if 'filter' in config:
        conditions = [f"{k} = ?" for k in config['filter']]
        where = "WHERE " + " AND ".join(conditions)
        params = list(config['filter'].values())

    rows = conn.execute(
        f"SELECT DISTINCT {first_col} FROM {table} {where} ORDER BY {first_col}", params
    ).fetchall()
    conn.close()
    return jsonify([r[first_col] for r in rows])


@app.route('/api/process_list', methods=['GET'])
def get_process_list():
    part = request.args.get('part', '')
    conn = get_db()
    if part:
        rows = conn.execute(
            "SELECT DISTINCT name FROM processes WHERE part = ? ORDER BY id", (part,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT DISTINCT name FROM processes ORDER BY id").fetchall()
    conn.close()
    return jsonify([r['name'] for r in rows])


# ==================== 통계 API ====================

@app.route('/api/stats/summary', methods=['GET'])
def get_stats_summary():
    conn = get_db()

    total = conn.execute(
        'SELECT COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt FROM scrap_data'
    ).fetchone()

    today = datetime.now().strftime('%Y-%m-%d')
    today_row = conn.execute(
        "SELECT COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt FROM scrap_data WHERE date LIKE ?",
        (f'{today}%',)
    ).fetchone()

    conn.close()

    return jsonify({
        'total': {
            'count': total['cnt'],
            'quantity': total['qty'] or 0,
            'weight': round(total['wgt'] or 0, 2)
        },
        'today': {
            'count': today_row['cnt'] or 0,
            'quantity': today_row['qty'] or 0,
            'weight': round(today_row['wgt'] or 0, 2)
        }
    })


@app.route('/api/stats/by_reason', methods=['GET'])
def get_stats_by_reason():
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = '''
        SELECT scrap_reason, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data WHERE 1=1
    '''
    params = []
    if start_date:
        query += ' AND date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND date <= ?'
        params.append(end_date + ' 23:59:59')
    query += ' GROUP BY scrap_reason ORDER BY cnt DESC'

    conn = get_db()
    results = [{
        'reason': row['scrap_reason'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in conn.execute(query, params).fetchall()]
    conn.close()
    return jsonify(results)


@app.route('/api/stats/by_part', methods=['GET'])
def get_stats_by_part():
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = '''
        SELECT part, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data WHERE 1=1
    '''
    params = []
    if start_date:
        query += ' AND date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND date <= ?'
        params.append(end_date + ' 23:59:59')
    query += ' GROUP BY part ORDER BY cnt DESC'

    conn = get_db()
    results = [{
        'part': row['part'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in conn.execute(query, params).fetchall()]
    conn.close()
    return jsonify(results)


@app.route('/api/stats/by_process', methods=['GET'])
def get_stats_by_process():
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = '''
        SELECT process, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data WHERE 1=1
    '''
    params = []
    if start_date:
        query += ' AND date >= ?'
        params.append(start_date)
    if end_date:
        query += ' AND date <= ?'
        params.append(end_date + ' 23:59:59')
    query += ' GROUP BY process ORDER BY cnt DESC'

    conn = get_db()
    results = [{
        'process': row['process'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in conn.execute(query, params).fetchall()]
    conn.close()
    return jsonify(results)


@app.route('/api/stats/daily', methods=['GET'])
def get_stats_daily():
    days = request.args.get('days', 30, type=int)

    conn = get_db()
    results = [{
        'date': row['day'],
        'count': row['cnt'],
        'quantity': row['qty'] or 0,
        'weight': round(row['wgt'] or 0, 2)
    } for row in conn.execute('''
        SELECT DATE(date) as day, COUNT(*) as cnt, SUM(quantity) as qty, SUM(weight) as wgt
        FROM scrap_data
        WHERE date >= DATE('now', ? || ' days')
        GROUP BY DATE(date)
        ORDER BY day DESC
    ''', (f'-{days}',)).fetchall()]
    conn.close()
    return jsonify(results)


# ==================== Excel 다운로드 API ====================

@app.route('/api/export_excel', methods=['GET'])
def export_excel():
    """폐기 기록을 Excel 파일로 다운로드 (4시트: Part별 × 사유별)"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    if not start_date or not end_date:
        return jsonify({'success': False, 'message': '시작일과 종료일을 입력해주세요.'}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl 라이브러리가 필요합니다.'}), 500

    conn = get_db()
    all_rows = conn.execute("""
        SELECT unique_id, date, part, department, process, machine, person,
               tmno, product_name, scrap_reason, quantity, weight, remark,
               defect_category, defect_process
        FROM scrap_data
        WHERE date >= ? AND date <= ?
        ORDER BY date DESC
    """, (start_date, end_date + ' 23:59:59')).fetchall()
    conn.close()

    main_reasons = ('공정불량', '셋팅불량')

    # 4개 시트 데이터 분류
    sheets_config = [
        ('1Part 공정·셋팅불량', [r for r in all_rows if r['part'] == '1Part' and r['scrap_reason'] in main_reasons]),
        ('2Part 공정·셋팅불량', [r for r in all_rows if r['part'] == '2Part' and r['scrap_reason'] in main_reasons]),
        ('1Part 기타폐기', [r for r in all_rows if r['part'] == '1Part' and r['scrap_reason'] not in main_reasons]),
        ('2Part 기타폐기', [r for r in all_rows if r['part'] == '2Part' and r['scrap_reason'] not in main_reasons]),
    ]

    # 스타일 정의
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = ['ID', '날짜', 'Part', '부서', '공정', '설비명', '폐기자',
               'TM-NO', '품명', '폐기사유', '수량', '중량(kg)', '비고',
               '불량해당', '해당공정']
    col_widths = [18, 20, 8, 10, 10, 15, 10, 15, 20, 15, 8, 12, 20, 10, 12]

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    for sheet_title, rows in sheets_config:
        ws = wb.create_sheet(title=sheet_title)
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row in rows:
            ws.append([
                row['unique_id'], row['date'], row['part'], row['department'],
                row['process'], row['machine'], row['person'], row['tmno'],
                row['product_name'], row['scrap_reason'], row['quantity'],
                row['weight'], row['remark'],
                row['defect_category'] or '', row['defect_process'] or ''
            ])

        for r in range(2, ws.max_row + 1):
            for cell in ws[r]:
                cell.alignment = data_align
                cell.border = thin_border

        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"scrap_data_{start_date}_{end_date}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== 서버 시작 ====================

if __name__ == '__main__':
    init_db()
    migrate_from_excel()
    print("=" * 50)
    print("폐기불량 관리시스템 서버 시작")
    print("http://localhost:5001 에서 접속하세요")
    print(f"SQLite DB: {DB_FILE}")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5001, debug=True)
